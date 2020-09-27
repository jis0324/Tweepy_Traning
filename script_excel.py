
import os
import tweepy
import openpyxl
import json

base_dir = os.path.dirname(os.path.abspath(__file__))
output_xlsx_path = base_dir + '/result.xlsx'

consumer_key = ""
consumer_secret = ""
access_token = ""
access_token_secret = ""

auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)

api = tweepy.API(auth)

class Profile:
  # init
  def __init__(self):
    self.users_list = list()
    self.get_users_list()

  # get users list
  def get_users_list(self):
    users_list_file = "{}/userlist.txt".format(base_dir)
    if os.path.isfile(users_list_file):
      with open(users_list_file, 'r') as users_list_txt:
        self.users_list=[row.rstrip('\n') for row in users_list_txt]
    else:
      print("----- Not Fount userlists.txt File -----")

  # create result.xlsx
  def create_xlsx(self):
    # Create a Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lessons"

    # write fieldnames
    ws.cell(row=1, column=1).value = "user_id"
    ws.cell(row=1, column=2).value = "status_id"
    ws.cell(row=1, column=3).value = "created_at"
    ws.cell(row=1, column=4).value = "screen_name"
    ws.cell(row=1, column=5).value = "name"
    ws.cell(row=1, column=6).value = "location"
    ws.cell(row=1, column=7).value = "description"
    ws.cell(row=1, column=8).value = "url"
    ws.cell(row=1, column=9).value = "protected"
    ws.cell(row=1, column=10).value = "followers_count"
    ws.cell(row=1, column=11).value = "friends_count"
    ws.cell(row=1, column=12).value = "listed_count"
    ws.cell(row=1, column=13).value = "statuses_count"
    ws.cell(row=1, column=14).value = "favourites_count"
    ws.cell(row=1, column=15).value = "account_created_at"
    ws.cell(row=1, column=16).value = "verified"
    ws.cell(row=1, column=17).value = "profile_url"
    ws.cell(row=1, column=18).value = "profile_expanded_url"
    ws.cell(row=1, column=19).value = "account_lang"
    ws.cell(row=1, column=20).value = "profile_banner_url"
    ws.cell(row=1, column=21).value = "profile_background_url"
    ws.cell(row=1, column=22).value = "profile_image_url"
    wb.save(output_xlsx_path)

  # insert profile to xlsx
  def insert_to_xlsx(self, profile_dict):
    file_exist = os.path.isfile(output_xlsx_path)
    if not file_exist:
      # create new result.xlsx
      self.create_xlsx()

    # excel load
    wb_obj = openpyxl.load_workbook(output_xlsx_path)

    # from the active attribute
    sheet_obj = wb_obj.active

    # get max column count
    max_row = sheet_obj.max_row
    row_index = max_row + 1

    sheet_obj.cell(row=row_index, column=1).value = profile_dict["user_id"]
    sheet_obj.cell(row=row_index, column=2).value = profile_dict["status_id"]
    sheet_obj.cell(row=row_index, column=3).value = profile_dict["created_at"]
    sheet_obj.cell(row=row_index, column=4).value = profile_dict["screen_name"]
    sheet_obj.cell(row=row_index, column=5).value = profile_dict["name"]
    sheet_obj.cell(row=row_index, column=6).value = profile_dict["location"]
    sheet_obj.cell(row=row_index, column=7).value = profile_dict["description"]
    sheet_obj.cell(row=row_index, column=8).value = profile_dict["url"]
    sheet_obj.cell(row=row_index, column=9).value = profile_dict["protected"]
    sheet_obj.cell(row=row_index, column=10).value = profile_dict["followers_count"]
    sheet_obj.cell(row=row_index, column=11).value = profile_dict["friends_count"]
    sheet_obj.cell(row=row_index, column=12).value = profile_dict["listed_count"]
    sheet_obj.cell(row=row_index, column=13).value = profile_dict["statuses_count"]
    sheet_obj.cell(row=row_index, column=14).value = profile_dict["favourites_count"]
    sheet_obj.cell(row=row_index, column=15).value = profile_dict["account_created_at"]
    sheet_obj.cell(row=row_index, column=16).value = profile_dict["verified"]
    sheet_obj.cell(row=row_index, column=17).value = profile_dict["profile_url"]
    sheet_obj.cell(row=row_index, column=18).value = profile_dict["profile_expanded_url"]
    sheet_obj.cell(row=row_index, column=19).value = profile_dict["account_lang"]
    sheet_obj.cell(row=row_index, column=20).value = profile_dict["profile_banner_url"]
    sheet_obj.cell(row=row_index, column=21).value = profile_dict["profile_background_url"]
    sheet_obj.cell(row=row_index, column=22).value = profile_dict["profile_image_url"]

    wb_obj.save(output_xlsx_path)

  # get profile of user
  def get_user_profile(self, username):
    user = api.get_user(username)

    profile_dict = dict()
    profile_dict["user_id"] = user.id_str
    profile_dict["status_id"] = ""
    profile_dict["created_at"] = str(user.created_at)
    profile_dict["screen_name"] = user.screen_name
    profile_dict["name"] = user.name
    profile_dict["location"] = user.location
    profile_dict["description"] = user.description
    profile_dict["url"] = user.url
    profile_dict["protected"] = user.protected
    profile_dict["followers_count"] = user.followers_count
    profile_dict["friends_count"] = user.followers_count
    profile_dict["listed_count"] = user.listed_count
    profile_dict["statuses_count"] = user.statuses_count
    profile_dict["favourites_count"] = user.favourites_count
    profile_dict["account_created_at"] = str(user.created_at)
    profile_dict["verified"] = user.verified
    profile_dict["profile_url"] = ""
    profile_dict["profile_expanded_url"] = ""
    if "url" in user.entities:
      if "urls" in user.entities["url"] and user.entities["url"]["urls"]:
        if "url" in user.entities["url"]["urls"][0]:
          profile_dict["profile_url"] = user.entities["url"]["urls"][0]["url"]
        if "expanded_url" in user.entities["url"]["urls"][0]:
          profile_dict["profile_expanded_url"] = user.entities["url"]["urls"][0]["expanded_url"]

    profile_dict["account_lang"] = user.lang
    profile_dict["profile_banner_url"] = user.profile_banner_url
    profile_dict["profile_background_url"] = user.profile_background_image_url
    profile_dict["profile_image_url"] = user.profile_image_url

    tweets = api.user_timeline(screen_name=username)
    for tweet in tweets:
      print(tweet.id)
      print(tweet.created_at)

    print(json.dumps(profile_dict, indent=2))
    return profile_dict

  # main
  def main(self):
    if self.users_list:
      for index, username in enumerate(self.users_list):
        print("{}th User / Total {} Users : {}".format(index+1, len(self.users_list), username))
        profile_dict = self.get_user_profile(username)
        self.insert_to_xlsx(profile_dict)

    else:
      print("----- Empty Users List, Please Fill Out Users and Try again. -----")
      return

if __name__ == '__main__':
  # delete output xlsx
  if os.path.isfile(output_xlsx_path):
    os.remove(output_xlsx_path)

  profile = Profile()
  profile.main()